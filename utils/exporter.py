from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib import colors
from reportlab.lib.units import mm


def export_to_excel(documents, path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Registru"

    headers = [
        "Nr. Înregistrare", "Tip", "Data", "Tip document",
        "Număr document", "Emitent/Destinatar", "Descriere", "Observații", "Atașament",
    ]

    header_fill = PatternFill("solid", fgColor="2563EB")
    header_font = Font(bold=True, color="FFFFFF", name="Calibri", size=11)

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for d in documents:
        ws.append([
            d["id"],
            d["tip_inout"],
            d["data"],
            d["tip_document"],
            d["numar_document"],
            d["emitent_destinatar"],
            d["descriere"],
            d["observatii"],
            d["attachment_path"] or "",
        ])

    # Auto-width
    col_widths = [8, 10, 12, 14, 18, 26, 36, 26, 30]
    for col, width in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(1, col).column_letter].width = width

    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"

    wb.save(path)


def export_to_pdf(documents, path):
    doc = SimpleDocTemplate(
        path,
        pagesize=landscape(A4),
        rightMargin=15 * mm,
        leftMargin=15 * mm,
        topMargin=20 * mm,
        bottomMargin=15 * mm,
    )

    styles = getSampleStyleSheet()
    cell_style = ParagraphStyle(
        "cell",
        parent=styles["Normal"],
        fontSize=8,
        leading=11,
        wordWrap="CJK",
    )
    header_style = ParagraphStyle(
        "header",
        parent=styles["Normal"],
        fontSize=8,
        leading=11,
        textColor=colors.white,
        fontName="Helvetica-Bold",
    )

    elements = []

    title_style = ParagraphStyle(
        "title",
        parent=styles["Title"],
        fontSize=16,
        textColor=colors.HexColor("#1e293b"),
        spaceAfter=4,
    )
    elements.append(Paragraph("Registru Intrări–Ieșiri", title_style))
    elements.append(Spacer(1, 8 * mm))

    col_labels = [
        "Nr.", "Tip", "Data", "Tip doc.",
        "Nr. document", "Emitent/Destinatar", "Descriere", "Observații",
    ]

    data = [[Paragraph(h, header_style) for h in col_labels]]

    for d in documents:
        data.append([
            Paragraph(str(d["id"]), cell_style),
            Paragraph(d["tip_inout"], cell_style),
            Paragraph(d["data"], cell_style),
            Paragraph(d["tip_document"], cell_style),
            Paragraph(d["numar_document"], cell_style),
            Paragraph(d["emitent_destinatar"], cell_style),
            Paragraph(d["descriere"] or "", cell_style),
            Paragraph(d["observatii"] or "", cell_style),
        ])

    # Column widths (landscape A4 usable ~257mm)
    col_widths_mm = [12, 18, 22, 22, 40, 50, 55, 38]
    col_widths_pts = [w * mm for w in col_widths_mm]

    table = Table(data, colWidths=col_widths_pts, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2563eb")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8fafc")]),
        ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#e2e8f0")),
        ("LINEBELOW", (0, 0), (-1, 0), 1.5, colors.HexColor("#1d4ed8")),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ("LEFTPADDING", (0, 0), (-1, -1), 6),
        ("RIGHTPADDING", (0, 0), (-1, -1), 6),
    ]))

    elements.append(table)
    doc.build(elements)
